# -*- coding: utf-8 -*-
"""
Created on Thu Jun 15 14:34:50 2017

@author: ndoannguyen
"""

from openpyxl import load_workbook
import pandas, string

def do_task_1(filename, bilan_writer, COL1, COL2, SHEET_NAME):
    """    
    Task 1:     Copy column "M" and column "D" (from row 2) of sheet "Donnees_bilan" of "file1.xls"
                Then paste it to column COL1 and COL2 (COL1, COL2 stand for "A", "B", "C"...) (from row 5) of sheet SHEET_NAME (SHEET_NAME stands for "vierge", "1000H", "3000H")    
    """    
    book = load_workbook(filename, data_only=True)
    sheet = book.get_sheet_by_name("Donnees_bilan")

    max_row = sheet.max_row
    for i in range(2, max_row + 1):
        #Here 2 because we copy data from the 2nd cell of the column
        if (sheet.cell(row = i, column = 13).value != None):
            bilan_writer.sheets[SHEET_NAME][COL1 + str(i + 3)] = sheet.cell(row = i, column = 13).value + 1 #13 is index of column M
        #Here i + 3 because we paste data to new sheet from the 5th cell of the column, and 5 - 2 = 3
        if (sheet.cell(row = i, column = 4).value != None):
            bilan_writer.sheets[SHEET_NAME][COL2 + str(i + 3)] = sheet.cell(row = i, column = 4).value + 1  #4 is index of column D

def do_task_2(filename, bilan_writer, ROW, SHEET_NAME):
    """    
    Task 2:     Copy row 2 of file "filename" (from cell D2)
                Then paste it to row ROW sheet SHEET_NAME of the bilan file (here SHEET_NAME may stand for 'Final_mechanical_data')
    """ 
    book = load_workbook(filename, data_only=True)
    sheet = book.get_sheet_by_name("Resultats")

    bilan_writer.sheets[SHEET_NAME]['A' + str(ROW)] = sheet.cell(row = 2, column = 1).value    
    
    UPPERCASE = string.uppercase
    for i in range(4, 26):          #That means from column D to column S
        bilan_writer.sheets[SHEET_NAME][UPPERCASE[i - 2] + str(ROW)] = sheet.cell(row = 2, column = i).value
        #Here i-2 because we want to paste column D of the partial file to column B of the final file, and so on 

    
bilan_workbook = load_workbook("file2.xlsx", data_only=True)
bilan_writer = pandas.ExcelWriter("file2.xlsx", engine='openpyxl') 
bilan_writer.book = bilan_workbook
bilan_writer.sheets = dict((ws.title, ws) for ws in bilan_workbook.worksheets)

"""
EXAMPLE, TO BE EDITED
"""

do_task_1('file1.xlsx', bilan_writer, 'A', 'B', 'vierge')
do_task_1('file1.xlsx', bilan_writer, 'E', 'F', 'vierge')
do_task_1('file1.xlsx', bilan_writer, 'C', 'D', '1000H')
do_task_1('file1.xlsx', bilan_writer, 'A', 'B', '2000H')
do_task_1('file1.xlsx', bilan_writer, 'C', 'D', '3000H')

do_task_2('file1.xlsx', bilan_writer, 30, 'Final_mechanical_data')
do_task_2('file1.xlsx', bilan_writer, 31, 'Final_mechanical_data')
do_task_2('file1.xlsx', bilan_writer, 32, 'Final_mechanical_data')
do_task_2('file1.xlsx', bilan_writer, 33, 'Final_mechanical_data')
do_task_2('file1.xlsx', bilan_writer, 35, 'Final_mechanical_data')

"""
END EXAMPLE
"""

bilan_workbook.save('file2_filled.xlsx')