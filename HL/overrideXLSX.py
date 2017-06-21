# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 16:42:53 2017

@author: ndoannguyen
"""

import pandas, string
from openpyxl import load_workbook

print list(string.uppercase)


book = load_workbook('G74_Rm_2.xlsx')
writer = pandas.ExcelWriter('G74_Rm_2.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# Add (100000, 200000, 300000) to row 40
writer.sheets['Feuil1']['A40'] = 100000
writer.sheets['Feuil1']['B40'] = 200000
writer.sheets['Feuil1']['C40'] = 300000

for v in range(1, 42):
    res = ""    
    for u in string.uppercase[:3]:
        res += "%s\t" % writer.sheets['Feuil1']["%s%s" % (u, str(v))].value
    print res

book.save('G74_Rm_2_new.xlsx')

print writer.sheets['Feuil1'].max_row