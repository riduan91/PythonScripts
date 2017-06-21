# -*- coding: utf-8 -*-
"""
Created on Fri Mar 17 11:58:22 2017

@author: ndoannguyen
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Mar 10 14:22:35 2017

@author: HL247224
"""
import numpy as np
from scipy.optimize import curve_fit
import pandas as pd
from openpyxl import load_workbook

def linear_model(X, a, b, c, d, e, f):
    x,y = X
    return a + b*x + c*y

def quadratic_model(X, a, b, c, d, e, f):
    x,y = X
    return a + b*x + c*y + d*x*y + e*x**2 + f*y**2
    
def exponential_model(X, a, b, c, d, e):
    x,y = X
    return a + b*np.exp(c*x) + d*np.exp(e*y)

def fit_and_save(func, inputfile, outputfile, newoutputfile, experiment_name):
    data = pd.read_excel(inputfile) #chergement du fichier Excel

    R_m = data['Rm'].values #selection de la variable, chargement dans une numpy array
    x = data['Temperature'].values
    y = data['Hydrogen'].values
    popt, pcov = curve_fit(func, (x,y), R_m)

    book = load_workbook(outputfile)
    writer = pd.ExcelWriter(outputfile, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    current_max_row = writer.sheets['Sheet1'].max_row
    new_row = current_max_row + 1
    
    writer.sheets['Sheet1']['A' + str(new_row)] = experiment_name
    column_index = 'B'
    for param in popt:
        writer.sheets['Sheet1'][column_index + str(new_row)] = param
        column_index = chr(ord(column_index) + 1)
        
    book.save(newoutputfile)
    
# MAIN PROGRAM
Experiment_Input_Files = ["G74_Rm_1.xlsx", "G74_Rm_2.xlsx"]
Experiment_Names = ["Experiment1", "Experiment2"]
Experiment_Output_Files = ["params.xlsx", "newparams.xlsx"]

for i in range(2):
    fit_and_save(quadratic_model, Experiment_Input_Files[i], Experiment_Output_Files[i], "newparams.xlsx", Experiment_Names[i])
